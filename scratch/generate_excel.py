import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Setup sheets
ws_instructions = wb.active
ws_instructions.title = "Instrucciones"
ws_prices = wb.create_sheet(title="Precios Unitarios")

# -------------------------------------------------------------
# TAB 1: INSTRUCCIONES
# -------------------------------------------------------------
ws_instructions.views.sheetView[0].showGridLines = True

# Styling tokens
font_title = Font(name="Segoe UI", size=16, bold=True, color="121820")
font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="667085")
font_heading = Font(name="Segoe UI", size=12, bold=True, color="121820")
font_body = Font(name="Segoe UI", size=10, color="20242a")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="20242a")
fill_header = PatternFill(start_color="121820", end_color="121820", fill_type="solid")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

# Write instructions content
ws_instructions["A2"] = "Planilla de Revisión de Precios Unitarios - HAM Estructuras Metálicas"
ws_instructions["A2"].font = font_title
ws_instructions["A3"] = "Proyecto: P-2026-0010 - Pérgola-Tendedero Casa Nura (Son Parc) · Cliente: Maximiliano Ventin"
ws_instructions["A3"].font = font_subtitle

ws_instructions["A5"] = "Instrucciones de Uso:"
ws_instructions["A5"].font = font_heading

instructions = [
    "1. Vaya a la pestaña 'Precios Unitarios'.",
    "2. Modifique los valores únicamente en la columna 'Nuevo Precio Unitario (EUR)' (Columna F, celda de color dorado).",
    "3. La columna 'Total Nuevo (EUR)' se recalculará automáticamente usando fórmulas de Excel.",
    "4. Al terminar de corregir los precios, guarde el archivo y envíelo de vuelta.",
    "5. Nuestra oficina técnica regenerará inmediatamente los archivos oficiales del presupuesto (HTML y PDF) con sus nuevos unitarios.",
    "Nota técnica: Las celdas de las columnas E y F están formateadas como moneda. Por favor ingrese solo números reales.",
]

for idx, text in enumerate(instructions, start=6):
    cell = ws_instructions[f"A{idx}"]
    cell.value = text
    cell.font = font_body
    if text.startswith("Nota técnica:"):
        cell.font = font_bold

# Adjust width of instructions column A
ws_instructions.column_dimensions["A"].width = 100

# -------------------------------------------------------------
# TAB 2: PRECIOS UNITARIOS
# -------------------------------------------------------------
ws_prices.views.sheetView[0].showGridLines = True

# Headers
headers = [
    "Categoría",
    "Concepto",
    "Cantidad",
    "Unidad",
    "Precio Unitario Actual (EUR)",
    "Nuevo Precio Unitario (EUR)",
    "Total Actual (EUR)",
    "Total Nuevo (EUR)",
    "Notas / Comentarios de Revisión"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws_prices.cell(row=2, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data items: Category, Concept, Qty, Unit, UnitPrice, Note
data_items = [
    # OPTION A
    ("Opción A (Acero Dúplex)", "Estructura base (Acero al carbono S275 bruto en kg, con 8% de merma)", 492.0, "kg", 1.60, "Incluye fustes, brazos y pletinas"),
    ("Opción A (Acero Dúplex)", "Tratamiento superficial Dúplex (Desengrase, granallado, galvanizado UNE-EN ISO 1461 + lacado al horno)", 455.0, "kg", 3.70, "Para ambiente C5 marino severo"),
    ("Opción A (Acero Dúplex)", "Cables Trellis e Inox AISI 316 (Cable Ø6 mm 1x19, stays superiores, tensores M8 y terminales oliva)", 264.0, "ml", 6.05, "Kit completo por metro de trellis"),
    ("Opción A (Acero Dúplex)", "Pernos cimentación M16 x 400 mm A4-80 inoxidables + Resina química bicomponente", 24.0, "ud", 25.00, "Anclaje químico estructural"),
    ("Opción A (Acero Dúplex)", "Tornillería pasante M12 fachada, camisas e Inox A4 con pletinas soporte", 20.0, "ud", 12.60, "Fijación a columnas de fachada"),
    ("Opción A (Acero Dúplex)", "Consumibles de taller específicos (soldadura, gas, discos de corte/repaso)", 1.0, "ud", 220.00, "Coste global consumibles"),
    ("Opción A (Acero Dúplex)", "Mano de obra Taller (Corte, armado, soldadura y repaso de estructura)", 60.0, "h", 30.00, "Oficial de herrería"),
    ("Opción A (Acero Dúplex)", "Mano de obra Montaje en Son Parc (Izado, aplomado, enhebrado de trellis y stays)", 48.0, "h", 30.00, "2 oficiales en obra"),
    ("Opción A (Acero Dúplex)", "Transporte marítimo Barcelona->Menorca y portes locales a pie de obra en Son Parc", 1.0, "ud", 350.00, "Coste de logística cerrado"),
    ("Opción A (Acero Dúplex)", "Oficina técnica (Planos de replanteo CNC, despiece tridimensional y coordinación)", 1.0, "ud", 400.00, "Coste ingeniería HAM"),
    
    # OPTION B
    ("Opción B (Inox AISI 316)", "Tubo fuste Inox AISI 316 Ø139,7x4 mm cepillado satinado (longitud con merma)", 22.1, "ml", 110.0, "Tarifa de fuste indicada por cliente"),
    ("Opción B (Inox AISI 316)", "Tubo brazos Inox AISI 316 60x60x4 mm en kg (con 8% de merma)", 91.8, "kg", 8.50, "Brazos en acero inoxidable"),
    ("Opción B (Inox AISI 316)", "Redondo macizo Inox AISI 316 Ø30 mm en kg (con 8% de merma)", 42.0, "kg", 8.50, "Cartelas de refuerzo en inoxidable"),
    ("Opción B (Inox AISI 316)", "Chapas base, rigidizadores y accesorios Inox AISI 316 en chapa 8/12mm (con mermas)", 54.3, "kg", 9.50, "Corte láser de chapas inox"),
    ("Opción B (Inox AISI 316)", "Tratamiento especial satinado fustes + Pintura de poliuretano al horno para brazos", 1.0, "ud", 975.00, "Acabados estéticos combinados"),
    ("Opción B (Inox AISI 316)", "Cables Trellis e Inox AISI 316 (Cable Ø6 mm 1x19, stays superiores, tensores M8 y terminales oliva)", 264.0, "ml", 6.05, "Kit completo por metro de trellis"),
    ("Opción B (Inox AISI 316)", "Pernos cimentación M16 x 400 mm A4-80 inoxidables + Resina química bicomponente", 24.0, "ud", 25.00, "Anclaje químico estructural"),
    ("Opción B (Inox AISI 316)", "Tornillería pasante M12 fachada, camisas e Inox A4 con pletinas soporte", 20.0, "ud", 12.60, "Fijación a columnas de fachada"),
    ("Opción B (Inox AISI 316)", "Consumibles de taller específicos (discos especiales inox, gas argón, hilo TIG)", 1.0, "ud", 380.00, "Coste global consumibles inox"),
    ("Opción B (Inox AISI 316)", "Mano de obra Taller (Corte, armado, soldadura TIG y pulido/satinado de inox)", 104.0, "h", 35.00, "Herrero especialista en acero inoxidable"),
    ("Opción B (Inox AISI 316)", "Mano de obra Montaje en Son Parc (Izado, aplomado, enhebrado de trellis y stays)", 48.0, "h", 30.00, "2 oficiales en obra"),
    ("Opción B (Inox AISI 316)", "Transporte marítimo Barcelona->Menorca y portes locales a pie de obra en Son Parc", 1.0, "ud", 350.00, "Coste de logística cerrado"),
    ("Opción B (Inox AISI 316)", "Oficina técnica (Planos de replanteo CNC, despiece tridimensional y coordinación)", 1.0, "ud", 500.00, "Coste ingeniería HAM Inox"),

    # CIVIL WORK
    ("Obra Civil (Opcional)", "Excavación de pozos de cimentación de 50x50x60 cm en terreno de Son Parc", 6.0, "ud", 58.33, "Pozos de anclaje para fustes"),
    ("Obra Civil (Opcional)", "Suministro de hormigón estructural HA-25/B/20/IIa, colocación, aditivos y transporte", 1.0, "m³", 220.00, "Vertido directo en pozos"),
    ("Obra Civil (Opcional)", "Armadura de reparto en acero corrugado B500S en parrillas inferiores de reparto", 1.0, "ud", 150.00, "Refuerzo estructural zapata"),
    ("Obra Civil (Opcional)", "Mano de obra replanteo, colocación de plantillas de precisión y vertido de hormigón", 16.0, "h", 30.00, "2 oficiales para cimientos"),
]

# Borders
thin_border = Border(
    left=Side(style='thin', color='d9e0e8'),
    right=Side(style='thin', color='d9e0e8'),
    top=Side(style='thin', color='d9e0e8'),
    bottom=Side(style='thin', color='d9e0e8')
)

fill_op_a = PatternFill(start_color="F6F8FB", end_color="F6F8FB", fill_type="solid")
fill_op_b = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_civil = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_edit = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # light gold/amber

# Write data items
for idx, item in enumerate(data_items, start=3):
    cat, concept, qty, unit, price, note = item
    
    ws_prices.cell(row=idx, column=1, value=cat).font = font_bold
    ws_prices.cell(row=idx, column=2, value=concept).font = font_body
    ws_prices.cell(row=idx, column=3, value=qty).font = font_body
    ws_prices.cell(row=idx, column=4, value=unit).font = font_body
    
    # Prices
    ws_prices.cell(row=idx, column=5, value=price).font = font_body
    ws_prices.cell(row=idx, column=5).number_format = '#,##0.00 €'
    
    # Editable prices
    edit_cell = ws_prices.cell(row=idx, column=6, value=price)
    edit_cell.font = font_bold
    edit_cell.fill = fill_edit
    edit_cell.number_format = '#,##0.00 €'
    
    # Formulas
    formula_act = f"=C{idx}*E{idx}"
    formula_new = f"=C{idx}*F{idx}"
    
    act_cell = ws_prices.cell(row=idx, column=7, value=formula_act)
    act_cell.font = font_body
    act_cell.number_format = '#,##0.00 €'
    
    new_cell = ws_prices.cell(row=idx, column=8, value=formula_new)
    new_cell.font = font_bold
    new_cell.number_format = '#,##0.00 €'
    
    # Notes
    ws_prices.cell(row=idx, column=9, value=note).font = font_body
    
    # Select fill based on category
    cat_fill = fill_op_a if "Opción A" in cat else (fill_civil if "Obra Civil" in cat else fill_op_b)
    for c in range(1, 10):
        cell = ws_prices.cell(row=idx, column=c)
        cell.border = thin_border
        if c != 6:  # don't overwrite edit cell fill
            cell.fill = cat_fill

# Add Totals row for categories
total_row_idx = len(data_items) + 4

# Op A Totals row
ws_prices.cell(row=total_row_idx, column=2, value="SUBTOTAL TECNICO BASE OPCION A").font = font_bold
ws_prices.cell(row=total_row_idx, column=7, value=f"=SUM(G3:G12)").font = font_bold
ws_prices.cell(row=total_row_idx, column=7).number_format = '#,##0.00 €'
ws_prices.cell(row=total_row_idx, column=8, value=f"=SUM(H3:H12)").font = font_bold
ws_prices.cell(row=total_row_idx, column=8).number_format = '#,##0.00 €'

# Op B Totals row
ws_prices.cell(row=total_row_idx+1, column=2, value="SUBTOTAL TECNICO BASE OPCION B").font = font_bold
ws_prices.cell(row=total_row_idx+1, column=7, value=f"=SUM(G13:G25)").font = font_bold
ws_prices.cell(row=total_row_idx+1, column=7).number_format = '#,##0.00 €'
ws_prices.cell(row=total_row_idx+1, column=8, value=f"=SUM(H13:H25)").font = font_bold
ws_prices.cell(row=total_row_idx+1, column=8).number_format = '#,##0.00 €'

# Civil Totals row
ws_prices.cell(row=total_row_idx+2, column=2, value="SUBTOTAL TECNICO BASE OBRA CIVIL").font = font_bold
ws_prices.cell(row=total_row_idx+2, column=7, value=f"=SUM(G26:G29)").font = font_bold
ws_prices.cell(row=total_row_idx+2, column=7).number_format = '#,##0.00 €'
ws_prices.cell(row=total_row_idx+2, column=8, value=f"=SUM(H26:H29)").font = font_bold
ws_prices.cell(row=total_row_idx+2, column=8).number_format = '#,##0.00 €'

border_total = Border(top=Side(style='double', color='121820'), bottom=Side(style='double', color='121820'))
for r in range(total_row_idx, total_row_idx+3):
    for c in range(1, 10):
        cell = ws_prices.cell(row=r, column=c)
        cell.font = font_bold
        cell.border = border_total
        cell.fill = fill_civil

# Auto-fit columns
for col in ws_prices.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    # Specific widths for headers to avoid extreme sizes
    if col_letter == 'A':
         ws_prices.column_dimensions[col_letter].width = 24
    elif col_letter == 'B':
         ws_prices.column_dimensions[col_letter].width = 50
    elif col_letter == 'E' or col_letter == 'F' or col_letter == 'G' or col_letter == 'H':
         ws_prices.column_dimensions[col_letter].width = 28
    else:
         ws_prices.column_dimensions[col_letter].width = max(max_len + 3, 10)

ws_prices.row_dimensions[2].height = 28

# Save workbook
output_path = "c:\\Users\\pablo\\Documents\\Presupuestador\\presupuestos\\P-2026-0010-pergola-tendedero-maximiliano-ventin\\precios-unitarios-pergola.xlsx"
wb.save(output_path)
print(f"Workbook successfully saved to: {output_path}")
